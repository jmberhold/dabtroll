#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET3_NAME = "sheet3_qwen3_5_MissionEngine_Review"
ERROR_NOTES_RE = re.compile(
    r"(cuda|unspecified launch failure|runtimeerror|exception|oom|out of memory|error)",
    re.IGNORECASE,
)
CUDA_RE = re.compile(r"(cuda|unspecified launch failure|cudaerrorlaunchfailure|cublas|device-side assert)", re.IGNORECASE)
EPISODE_RE = re.compile(r"_episode_(\d+)_of_21_")
REQUIRED_FILES = [
    "episode_summary.json",
    "bt.json",
    "status_window_manifest.jsonl",
    "human_rater_evaluation.xlsx",
]


def _discover_eligible(root: Path) -> list[Path]:
    episodes: list[Path] = []
    for ep in sorted(root.glob("*episode_*_of_21_*")):
        m = EPISODE_RE.search(ep.name)
        if not m:
            continue
        idx = int(m.group(1))
        if idx < 2 or idx > 21:
            continue
        if all((ep / r).exists() for r in REQUIRED_FILES):
            episodes.append(ep)
    return episodes


def _latest_qwen_output_dir(ep: Path) -> Path | None:
    candidates = sorted(ep.glob("qwen_3_5_*"))
    return candidates[-1] if candidates else None


def _audit_qwen_output(ep: Path, min_response_ok_ratio: float) -> dict[str, Any]:
    out: dict[str, Any] = {
        "qwen_output_dir": "",
        "mission_response_rows": 0,
        "mission_response_ok_rows": 0,
        "mission_response_error_rows": 0,
        "mission_cuda_error_rows": 0,
        "mission_empty_text_rows": 0,
        "mission_response_ok_ratio": 0.0,
        "qwen_manifest_rows": 0,
        "qwen_manifest_blank_status_rows": 0,
        "qwen_manifest_response_not_ok_rows": 0,
        "qwen_manifest_cuda_note_rows": 0,
        "qwen_quality_ok": False,
        "qwen_quality_reason": "",
    }

    qdir = _latest_qwen_output_dir(ep)
    if qdir is None:
        out["qwen_quality_reason"] = "missing_qwen_output_dir"
        return out

    out["qwen_output_dir"] = str(qdir)

    mission_path = qdir / "missionengine.jsonl"
    if not mission_path.exists():
        out["qwen_quality_reason"] = "missing_missionengine_jsonl"
        return out

    try:
        lines = mission_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception as exc:  # pragma: no cover
        out["qwen_quality_reason"] = f"missionengine_read_failed: {exc}"
        return out

    for line in lines:
        text = line.strip()
        if not text:
            continue
        try:
            row = json.loads(text)
        except Exception:
            continue
        if str(row.get("direction", "") or "") != "response":
            continue

        out["mission_response_rows"] += 1
        response_ok = bool(row.get("response_ok", False))
        response_error = str(row.get("response_error", "") or "")
        response_text = str(row.get("response_text", "") or "")
        response_blob = row.get("response") if isinstance(row.get("response"), dict) else {}
        response_error_blob = str(response_blob.get("error", "") or "")

        if response_ok:
            out["mission_response_ok_rows"] += 1
        else:
            out["mission_response_error_rows"] += 1

        if (not response_ok) and (not response_text.strip()):
            out["mission_empty_text_rows"] += 1

        haystack = "\n".join([response_error, response_error_blob, response_text])
        if CUDA_RE.search(haystack):
            out["mission_cuda_error_rows"] += 1

    if out["mission_response_rows"] > 0:
        out["mission_response_ok_ratio"] = (
            float(out["mission_response_ok_rows"]) / float(out["mission_response_rows"])
        )

    manifest_path = qdir / "status_window_manifest.jsonl"
    if manifest_path.exists():
        for line in manifest_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            text = line.strip()
            if not text:
                continue
            try:
                row = json.loads(text)
            except Exception:
                continue
            out["qwen_manifest_rows"] += 1

            status = str(row.get("qwen35_status", "") or "").strip().lower()
            response_ok = bool(row.get("qwen35_response_ok", False))
            notes = str(row.get("qwen35_notes", "") or "")
            resp_text = str(row.get("qwen35_response_text", "") or "")

            if not status:
                out["qwen_manifest_blank_status_rows"] += 1
            if not response_ok:
                out["qwen_manifest_response_not_ok_rows"] += 1
            if CUDA_RE.search("\n".join([notes, resp_text])):
                out["qwen_manifest_cuda_note_rows"] += 1

    reasons: list[str] = []
    if out["mission_response_rows"] == 0:
        reasons.append("no_mission_responses")
    if out["mission_response_ok_ratio"] < float(min_response_ok_ratio):
        reasons.append(f"low_response_ok_ratio={out['mission_response_ok_ratio']:.3f}")
    if out["mission_cuda_error_rows"] > 0:
        reasons.append(f"mission_cuda_rows={out['mission_cuda_error_rows']}")
    if out["mission_response_error_rows"] > 0:
        reasons.append(f"mission_error_rows={out['mission_response_error_rows']}")
    if out["qwen_manifest_response_not_ok_rows"] > 0:
        reasons.append(f"manifest_not_ok_rows={out['qwen_manifest_response_not_ok_rows']}")
    if out["qwen_manifest_cuda_note_rows"] > 0:
        reasons.append(f"manifest_cuda_rows={out['qwen_manifest_cuda_note_rows']}")
    if out["qwen_manifest_blank_status_rows"] > 0:
        reasons.append(f"manifest_blank_status={out['qwen_manifest_blank_status_rows']}")

    out["qwen_quality_ok"] = len(reasons) == 0
    out["qwen_quality_reason"] = "clean_qwen_output" if not reasons else ",".join(reasons)
    return out


def _audit_episode(ep: Path, min_response_ok_ratio: float) -> dict[str, Any]:
    row: dict[str, Any] = {
        "episode_dir": str(ep),
        "keep": False,
        "reason": "",
        "sheet3_rows": 0,
        "error_rows": 0,
        "unparsed_rows": 0,
        "blank_status_rows": 0,
        "error_note_rows": 0,
    }

    row.update(_audit_qwen_output(ep, min_response_ok_ratio=min_response_ok_ratio))

    wb_path = ep / "human_rater_evaluation.xlsx"
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Title is more than 31 characters.*",
                category=UserWarning,
            )
            wb = load_workbook(wb_path, data_only=True)
    except Exception as exc:  # pragma: no cover
        row["reason"] = f"workbook_open_failed: {exc}"
        return row

    if SHEET3_NAME not in wb.sheetnames:
        row["reason"] = "missing_sheet3"
        return row

    ws = wb[SHEET3_NAME]
    if ws.max_row < 9:
        row["reason"] = "sheet3_no_data_rows"
        return row

    for r in range(9, ws.max_row + 1):
        source = str(ws.cell(r, 2).value or "").strip()
        if not source:
            continue

        row["sheet3_rows"] += 1
        status = str(ws.cell(r, 10).value or "").strip().lower()
        notes = str(ws.cell(r, 12).value or "").strip()

        if not status:
            row["blank_status_rows"] += 1
        if status == "error":
            row["error_rows"] += 1
        if status == "unparsed":
            row["unparsed_rows"] += 1
        if ERROR_NOTES_RE.search(notes):
            row["error_note_rows"] += 1

    if row["sheet3_rows"] == 0:
        row["reason"] = "sheet3_no_milestones"
        return row

    bad = (
        row["error_rows"] > 0
        or row["unparsed_rows"] > 0
        or row["blank_status_rows"] > 0
        or row["error_note_rows"] > 0
    )

    if bad:
        reasons: list[str] = []
        if row["error_rows"]:
            reasons.append(f"error_status={row['error_rows']}")
        if row["unparsed_rows"]:
            reasons.append(f"unparsed_status={row['unparsed_rows']}")
        if row["blank_status_rows"]:
            reasons.append(f"blank_status={row['blank_status_rows']}")
        if row["error_note_rows"]:
            reasons.append(f"error_notes={row['error_note_rows']}")
        row["reason"] = ",".join(reasons)
    else:
        row["reason"] = "clean_sheet3"

    if not bool(row.get("qwen_quality_ok", False)):
        if row["reason"]:
            row["reason"] = f"{row['reason']},{row.get('qwen_quality_reason','')}"
        else:
            row["reason"] = str(row.get("qwen_quality_reason", ""))

    row["keep"] = bool(row["reason"] == "clean_sheet3,clean_qwen_output" or row["reason"] == "clean_sheet3") and bool(row.get("qwen_quality_ok", False))
    if row["keep"]:
        row["reason"] = "clean_sheet3_and_qwen_output"

    return row


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Audit qwen3.5 Sheet3 quality for episodes 2-21")
    ap.add_argument("--logs-root", default="/home/mark/dabtroll/data/logs")
    ap.add_argument(
        "--episode-dir",
        action="append",
        default=[],
        help="Optional specific episode dir(s). If omitted, auto-discovers all eligible episode 2-21 dirs.",
    )
    ap.add_argument("--audit-json-out", default="")
    ap.add_argument("--keep-manifest-out", default="")
    ap.add_argument("--rerun-manifest-out", default="")
    ap.add_argument(
        "--state-jsonl",
        default="",
        help="Optional replay state jsonl. If provided with --state-only-ok, audits only successful episode_dir entries from this state.",
    )
    ap.add_argument(
        "--state-only-ok",
        action="store_true",
        help="When used with --state-jsonl, include only rows with ok=true.",
    )
    ap.add_argument(
        "--min-response-ok-ratio",
        type=float,
        default=0.95,
        help="Minimum required ratio of response_ok=true rows in missionengine.jsonl latest qwen output.",
    )
    return ap.parse_args()


def main() -> None:
    args = _parse_args()
    root = Path(args.logs_root).expanduser().resolve()

    if args.episode_dir:
        episodes = [Path(p).expanduser().resolve() for p in args.episode_dir]
    elif args.state_jsonl:
        seen: set[str] = set()
        episodes = []
        state_path = Path(args.state_jsonl).expanduser().resolve()
        for line in state_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            text = line.strip()
            if not text:
                continue
            try:
                row = json.loads(text)
            except Exception:
                continue
            if args.state_only_ok and not bool(row.get("ok")):
                continue
            ep = str(row.get("episode_dir", "") or "").strip()
            if not ep or ep in seen:
                continue
            seen.add(ep)
            episodes.append(Path(ep).expanduser().resolve())
    else:
        episodes = _discover_eligible(root)

    rows = [_audit_episode(ep, min_response_ok_ratio=float(args.min_response_ok_ratio)) for ep in episodes]
    keep = [r["episode_dir"] for r in rows if r.get("keep")]
    rerun = [r["episode_dir"] for r in rows if not r.get("keep")]

    payload = {
        "total_eligible": len(episodes),
        "keep_count": len(keep),
        "rerun_count": len(rerun),
        "audit_rows": rows,
    }

    if args.audit_json_out:
        Path(args.audit_json_out).write_text(json.dumps(payload, indent=2), encoding="utf-8")
    if args.keep_manifest_out:
        Path(args.keep_manifest_out).write_text("\n".join(keep) + ("\n" if keep else ""), encoding="utf-8")
    if args.rerun_manifest_out:
        Path(args.rerun_manifest_out).write_text("\n".join(rerun) + ("\n" if rerun else ""), encoding="utf-8")

    print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
