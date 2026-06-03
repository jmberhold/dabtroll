from __future__ import annotations

"""Generate a 2-sheet Excel workbook for human BT and mission-engine review.

This script creates an .xlsx file with:
1) BT node review sheet (node-by-node independent human assessment)
2) Mission engine milestone review sheet (agreement/disagreement with model outputs)

Notes:
- Excel/openpyxl does not natively embed SVG in a reliable cross-platform way.
  The workbook includes a prominent hyperlink to bt.svg at the top of each sheet.
- If btstatus/missionengine logs exist, the workbook is pre-filled with machine
  suggestions to speed up rating while keeping reviewer fields editable.
"""

import argparse
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ModuleNotFoundError as exc:  # pragma: no cover - import guard for runtime env differences
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: python -m pip install openpyxl"
    ) from exc


HEADER_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text:
            continue
        try:
            obj = json.loads(text)
        except Exception:
            continue
        if isinstance(obj, dict):
            out.append(obj)
    return out


def _fmt_mmss_from_step(step: Optional[int], outer_step_seconds: float) -> str:
    if step is None:
        return ""
    total_seconds = max(0.0, float(step) * float(outer_step_seconds))
    minutes = int(total_seconds // 60)
    seconds = int(round(total_seconds - (minutes * 60)))
    if seconds == 60:
        minutes += 1
        seconds = 0
    return f"{minutes}:{seconds:02d}"


def _fmt_mmss_from_seconds(seconds: Optional[float]) -> str:
    if seconds is None:
        return ""
    total_seconds = max(0.0, float(seconds))
    minutes = int(total_seconds // 60)
    rem_seconds = int(round(total_seconds - (minutes * 60)))
    if rem_seconds == 60:
        minutes += 1
        rem_seconds = 0
    return f"{minutes}:{rem_seconds:02d}"


def _safe_video_duration_seconds(video_path: Path) -> Optional[float]:
    """Return video duration in seconds using available local backends."""
    if not video_path.exists():
        return None

    try:
        import cv2  # type: ignore

        cap = cv2.VideoCapture(str(video_path))
        if cap.isOpened():
            frame_count = float(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0.0)
            fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
            cap.release()
            if fps > 0.0 and frame_count >= 0.0:
                return frame_count / fps
        else:
            cap.release()
    except Exception:
        pass

    try:
        import av  # type: ignore

        with av.open(str(video_path)) as container:
            if container.duration is not None:
                return float(container.duration) / 1_000_000.0
    except Exception:
        pass

    return None


def _find_primary_video_path(episode_dir: Path) -> Optional[Path]:
    video_dir = episode_dir / "video"
    if not video_dir.exists():
        return None
    mp4s = sorted(video_dir.glob("**/*.mp4"))
    if not mp4s:
        return None
    return mp4s[0]


def _frame_token_from_paths(frame_paths: Any) -> str:
    if isinstance(frame_paths, list) and frame_paths:
        last = Path(str(frame_paths[-1])).stem
        if last.startswith("frame_"):
            return last
    return ""


def _frame_token_from_entry(entry: Dict[str, Any]) -> str:
    token = str(entry.get("current_frame", "") or "").strip()
    if token.startswith("frame_"):
        return token
    frame_paths = entry.get("frames", [])
    token = _frame_token_from_paths(frame_paths)
    if token:
        return token
    current_path = str(entry.get("current_frame_path", "") or "")
    stem = Path(current_path).stem
    if stem.startswith("frame_"):
        return stem
    step = entry.get("step")
    if isinstance(step, int) and step >= 0:
        return f"frame_{step:05d}"
    return ""


def _flatten_bt_nodes(root: Dict[str, Any]) -> List[Dict[str, Any]]:
    nodes: List[Dict[str, Any]] = []

    def walk(node: Dict[str, Any]) -> None:
        if not isinstance(node, dict):
            return
        node_type = str(node.get("type", "")).strip().lower()
        if node_type in {"action", "condition"}:
            details = node.get(node_type, {}) if isinstance(node.get(node_type), dict) else {}
            nodes.append(
                {
                    "id": str(node.get("id", "")),
                    "type": node_type,
                    "description": str(
                        details.get("description")
                        or node.get("description")
                        or ""
                    ),
                    "success_criteria": str(details.get("success_criteria") or ""),
                }
            )
        children = node.get("children", [])
        if isinstance(children, list):
            for child in children:
                if isinstance(child, dict):
                    walk(child)

    walk(root)
    return nodes


def _build_bt_completion_suggestions(
    btstatus_rows: List[Dict[str, Any]],
    outer_step_seconds: float,
) -> Dict[str, Dict[str, str]]:
    suggestions: Dict[str, Dict[str, str]] = {}
    sorted_rows = sorted(
        btstatus_rows,
        key=lambda r: (
            int(r.get("step", 10**9)) if isinstance(r.get("step"), int) else 10**9,
            float(r.get("ts", 0.0)) if isinstance(r.get("ts"), (int, float)) else 0.0,
        ),
    )
    for row in sorted_rows:
        node_id = str(row.get("node_id", "") or "")
        if not node_id:
            continue
        status = str(row.get("status", "") or "").strip().lower()
        if status != "complete":
            continue
        if node_id in suggestions:
            continue
        step = row.get("step") if isinstance(row.get("step"), int) else None
        suggestions[node_id] = {
            "machine_time": _fmt_mmss_from_step(step, outer_step_seconds),
            "machine_frame": _frame_token_from_paths(row.get("frames", [])),
            "machine_notes": str(row.get("notes", "") or ""),
            "status_eval_id": str(row.get("status_eval_id", "") or ""),
        }
    return suggestions


def _parse_status_from_response_text(text: str) -> Tuple[str, str, str]:
    if not text:
        return "", "", ""
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            status = str(parsed.get("status", "") or "")
            progress = str(parsed.get("progress_score", parsed.get("progress", "")) or "")
            notes = str(parsed.get("notes", "") or "")
            return status, progress, notes
    except Exception:
        pass
    return "", "", ""


def _build_mission_milestones(
    mission_rows: List[Dict[str, Any]],
    btstatus_rows: List[Dict[str, Any]],
    outer_step_seconds: float,
    video_time_scale: Optional[float],
) -> List[Dict[str, Any]]:
    milestones: List[Dict[str, Any]] = []

    def _video_time_from_step(step: Optional[int]) -> str:
        if step is None or video_time_scale is None:
            return ""
        machine_seconds = max(0.0, float(step) * float(outer_step_seconds))
        return _fmt_mmss_from_seconds(machine_seconds * float(video_time_scale))

    for row in mission_rows:
        direction = str(row.get("direction", "") or "")
        request_kind = str(row.get("request_kind", "") or "")
        if direction != "response":
            continue
        if "status_eval" not in request_kind:
            continue
        step = row.get("step") if isinstance(row.get("step"), int) else None
        response = row.get("response") if isinstance(row.get("response"), dict) else {}
        response_text = str(row.get("response_text", "") or response.get("text", "") or "")
        model_status, model_progress, model_notes = _parse_status_from_response_text(response_text)
        milestones.append(
            {
                "sort_step": step if step is not None else 10**9,
                "sort_ts": float(row.get("ts", 0.0)) if isinstance(row.get("ts"), (int, float)) else 0.0,
                "source": "mission_engine_response",
                "request_kind": request_kind,
                "status_eval_id": str(row.get("status_eval_id", "") or ""),
                "node_id": str(row.get("node_id", "") or ""),
                "node_type": str(row.get("node_type", "") or ""),
                "frame": _frame_token_from_entry(row),
                "time": _fmt_mmss_from_step(step, outer_step_seconds),
                "video_time": _video_time_from_step(step),
                "model_status": model_status,
                "model_progress": model_progress,
                "model_notes": model_notes,
                "model_latency_s": response.get("latency_s", row.get("response_latency_s", "")),
            }
        )

    for row in btstatus_rows:
        status = str(row.get("status", "") or "").strip().lower()
        if status != "complete":
            continue
        step = row.get("step") if isinstance(row.get("step"), int) else None
        milestones.append(
            {
                "sort_step": step if step is not None else 10**9,
                "sort_ts": float(row.get("ts", 0.0)) if isinstance(row.get("ts"), (int, float)) else 0.0,
                "source": "btstatus_complete_milestone",
                "request_kind": "node_completion",
                "status_eval_id": str(row.get("status_eval_id", "") or ""),
                "node_id": str(row.get("node_id", "") or ""),
                "node_type": str(row.get("node_type", "") or ""),
                "frame": _frame_token_from_paths(row.get("frames", [])),
                "time": _fmt_mmss_from_step(step, outer_step_seconds),
                "video_time": _video_time_from_step(step),
                "model_status": str(row.get("status", "") or ""),
                "model_progress": str(row.get("progress_score", "") or ""),
                "model_notes": str(row.get("notes", "") or ""),
                "model_latency_s": "",
            }
        )

    milestones.sort(key=lambda m: (m["sort_step"], m["sort_ts"], m["source"]))
    return milestones


def _style_header_row(ws, row_idx: int, columns: Iterable[int]) -> None:
    for col in columns:
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _style_data_grid(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c in (12, 18):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def _write_sheet1(
    wb: Workbook,
    summary: Dict[str, Any],
    bt_nodes: List[Dict[str, Any]],
    bt_suggestions: Dict[str, Dict[str, str]],
    bt_svg_path: Path,
) -> None:
    ws = wb.active
    ws.title = "Sheet1_BT_Node_Review"

    title = "Human Review Sheet 1: Independent Behavior Tree Node Evaluation"
    ws.merge_cells("A1:M1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    instructions = (
        "Instructions: Review the BT and episode video independently. For each action/condition node, enter the "
        "reviewer completion time in m:ss using video start 0:00 at frame_00000. Add completion frame token, "
        "confidence (1-5), and notes about ambiguity, failure modes, or evidence quality. Use machine suggestions "
        "only as reference. Record disagreements and rationale in notes."
    )
    ws.merge_cells("A2:M4")
    ws["A2"] = instructions
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A5:M5")
    ws["A5"] = f"BT SVG (open externally): {bt_svg_path}"
    ws["A5"].hyperlink = bt_svg_path.as_uri() if bt_svg_path.exists() else None
    ws["A5"].font = Font(color="0563C1", underline="single") if bt_svg_path.exists() else Font(bold=True)

    ws.merge_cells("A6:M6")
    ws["A6"] = (
        f"Episode: {summary.get('episode_id', '')} | Mission: {summary.get('mission_name', '')} | "
        f"Env: {summary.get('env_name', '')}"
    )

    headers = [
        "node_order",
        "node_id",
        "node_type",
        "node_description",
        "success_criteria",
        "reviewer_completion_time_m:ss",
        "reviewer_completion_frame",
        "machine_suggested_time_m:ss",
        "machine_suggested_frame",
        "reviewer_agrees_machine_YN_partial",
        "reviewer_confidence_1to5",
        "reviewer_notes",
        "status_eval_id_ref",
    ]
    header_row = 8
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=h)
    _style_header_row(ws, header_row, range(1, len(headers) + 1))

    row = header_row + 1
    for i, node in enumerate(bt_nodes, start=1):
        node_id = str(node.get("id", ""))
        sugg = bt_suggestions.get(node_id, {})
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=node_id)
        ws.cell(row=row, column=3, value=node.get("type", ""))
        ws.cell(row=row, column=4, value=node.get("description", ""))
        ws.cell(row=row, column=5, value=node.get("success_criteria", ""))
        ws.cell(row=row, column=8, value=sugg.get("machine_time", ""))
        ws.cell(row=row, column=9, value=sugg.get("machine_frame", ""))
        ws.cell(row=row, column=13, value=sugg.get("status_eval_id", ""))
        row += 1

    data_end = max(header_row + 1, row - 1)
    _style_data_grid(ws, header_row, data_end, 1, len(headers))

    widths = {
        1: 10,
        2: 28,
        3: 12,
        4: 38,
        5: 42,
        6: 24,
        7: 24,
        8: 24,
        9: 24,
        10: 26,
        11: 22,
        12: 40,
        13: 34,
    }
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    # Keep machine-suggested fields for audit traceability, but hide from raters on sheet 1.
    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True

    ws.freeze_panes = "A9"


def _write_sheet2(
    wb: Workbook,
    summary: Dict[str, Any],
    milestones: List[Dict[str, Any]],
    bt_svg_path: Path,
) -> None:
    ws = wb.create_sheet("Sheet2_MissionEngine_Review")

    ws.merge_cells("A1:R1")
    ws["A1"] = "Human Review Sheet 2: Mission Engine Response and Milestone Agreement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = TITLE_FILL

    instructions = (
        "Instructions: Review major milestones (especially node completion points). Video time starts at 0:00 for "
        "frame_00000. For each row, judge whether mission-engine assessment is correct, enter reviewer time/frame "
        "if different, and explain disagreement. Prioritize correctness of completion timing, status label, and evidence."
    )
    ws.merge_cells("A2:R4")
    ws["A2"] = instructions
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A5:R5")
    ws["A5"] = f"BT SVG (open externally): {bt_svg_path}"
    ws["A5"].hyperlink = bt_svg_path.as_uri() if bt_svg_path.exists() else None
    ws["A5"].font = Font(color="0563C1", underline="single") if bt_svg_path.exists() else Font(bold=True)

    ws.merge_cells("A6:R6")
    ws["A6"] = (
        f"Episode: {summary.get('episode_id', '')} | Mission: {summary.get('mission_name', '')} | "
        f"Outer step seconds: {summary.get('outer_step_seconds', '')}"
    )

    headers = [
        "milestone_idx",
        "source",
        "request_kind_or_event",
        "status_eval_id",
        "node_id",
        "node_type",
        "machine_frame",
        "machine_time_m:ss",
        "video_time_m:ss",
        "model_status",
        "model_progress",
        "model_notes",
        "model_latency_s",
        "reviewer_agree_YN_partial",
        "reviewer_time_m:ss",
        "reviewer_frame",
        "reviewer_status",
        "reviewer_notes",
        "evidence_refs",
    ]
    header_row = 8
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=h)
    _style_header_row(ws, header_row, range(1, len(headers) + 1))

    row = header_row + 1
    for i, m in enumerate(milestones, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=m.get("source", ""))
        ws.cell(row=row, column=3, value=m.get("request_kind", ""))
        ws.cell(row=row, column=4, value=m.get("status_eval_id", ""))
        ws.cell(row=row, column=5, value=m.get("node_id", ""))
        ws.cell(row=row, column=6, value=m.get("node_type", ""))
        ws.cell(row=row, column=7, value=m.get("frame", ""))
        ws.cell(row=row, column=8, value=m.get("time", ""))
        ws.cell(row=row, column=9, value=m.get("video_time", ""))
        ws.cell(row=row, column=10, value=m.get("model_status", ""))
        ws.cell(row=row, column=11, value=m.get("model_progress", ""))
        ws.cell(row=row, column=12, value=m.get("model_notes", ""))
        ws.cell(row=row, column=13, value=m.get("model_latency_s", ""))
        row += 1

    data_end = max(header_row + 1, row - 1)
    _style_data_grid(ws, header_row, data_end, 1, len(headers))

    widths = {
        1: 12,
        2: 26,
        3: 28,
        4: 34,
        5: 28,
        6: 12,
        7: 16,
        8: 18,
        9: 16,
        10: 16,
        11: 14,
        12: 42,
        13: 14,
        14: 24,
        15: 20,
        16: 18,
        17: 18,
        18: 40,
        19: 34,
    }
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    ws.freeze_panes = "A9"


def _resolve_episode_paths(episode_dir: Path) -> Dict[str, Path]:
    return {
        "summary": episode_dir / "episode_summary.json",
        "bt_json": episode_dir / "bt.json",
        "bt_svg": episode_dir / "bt.svg",
        "btstatus": episode_dir / "btstatus.jsonl",
        "missionengine": episode_dir / "missionengine.jsonl",
    }


def generate_workbook(episode_dir: Path, output_path: Path) -> Path:
    paths = _resolve_episode_paths(episode_dir)
    missing = [name for name, p in paths.items() if name in {"summary", "bt_json"} and not p.exists()]
    if missing:
        raise FileNotFoundError(
            "Missing required files in episode directory: " + ", ".join(missing)
        )

    summary = _read_json(paths["summary"])
    bt = _read_json(paths["bt_json"])
    bt_root = bt.get("root", {}) if isinstance(bt, dict) else {}
    bt_nodes = _flatten_bt_nodes(bt_root if isinstance(bt_root, dict) else {})

    btstatus_rows = _read_jsonl(paths["btstatus"])
    mission_rows = _read_jsonl(paths["missionengine"])

    outer_step_seconds = float(summary.get("outer_step_seconds", 0.4) or 0.4)
    steps_executed = int(summary.get("steps_executed", 0) or 0)
    machine_duration_seconds = max(0.0, float(steps_executed) * float(outer_step_seconds))

    video_duration_seconds: Optional[float] = None
    video_time_scale: Optional[float] = None
    video_path = _find_primary_video_path(episode_dir)
    if video_path is not None:
        video_duration_seconds = _safe_video_duration_seconds(video_path)
    if video_duration_seconds is not None and machine_duration_seconds > 0.0:
        video_time_scale = float(video_duration_seconds) / float(machine_duration_seconds)

    bt_suggestions = _build_bt_completion_suggestions(btstatus_rows, outer_step_seconds)
    milestones = _build_mission_milestones(
        mission_rows,
        btstatus_rows,
        outer_step_seconds,
        video_time_scale,
    )

    wb = Workbook()
    _write_sheet1(wb, summary, bt_nodes, bt_suggestions, paths["bt_svg"])
    _write_sheet2(wb, summary, milestones, paths["bt_svg"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _default_output(episode_dir: Path) -> Path:
    return episode_dir / "human_rater_evaluation.xlsx"


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate human-rater BT/mission-engine evaluation workbook")
    ap.add_argument(
        "--episode-dir",
        required=True,
        help="Path to one episode folder under data/logs/...",
    )
    ap.add_argument(
        "--output",
        default="",
        help="Output .xlsx path (default: <episode-dir>/human_rater_evaluation.xlsx)",
    )
    args = ap.parse_args()

    episode_dir = Path(args.episode_dir).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else _default_output(episode_dir)

    out = generate_workbook(episode_dir, output_path)
    print(json.dumps({"ok": True, "workbook_path": str(out)}, indent=2))


if __name__ == "__main__":
    main()
