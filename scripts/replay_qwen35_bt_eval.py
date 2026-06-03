from __future__ import annotations

"""Replay mission-engine status evaluation on existing episode frames using Qwen3.5.

This script does NOT run simulation. It reuses:
- existing BT (`bt.json`)
- existing frame windows (`status_window_manifest.jsonl`)

For each episode, it creates a subfolder:
  qwen_3_5_<timestamp>/
with:
- status_window_manifest.jsonl
- missionengine.jsonl
- pipeline_trace.jsonl

It also appends/overwrites only Sheet 3 in existing human workbook:
- Sheet name: sheet3_qwen3_5_MissionEngine_Review
"""

import argparse
import base64
import io
import json
import time
import warnings
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from PIL import Image

from dabtroll_bt_pipeline import (
    MissionEngineClient,
    PipelineConfig,
    _status_eval_every_steps,
    _status_window_frame_count,
)
from mission_engine import BehaviorTreeRunner
from prompts import status_eval_text

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: python -m pip install openpyxl"
    ) from exc


HEADER_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not path.exists():
        return rows
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text:
            continue
        try:
            obj = json.loads(text)
        except Exception:
            continue
        if isinstance(obj, dict):
            rows.append(obj)
    return rows


def _write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=True) + "\n")


def _append_jsonl(path: Path, row: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=True) + "\n")


def _encode_jpg_file(path: Path) -> str:
    img = Image.open(path).convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=92)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def _parse_status_json(text: str) -> Dict[str, Any]:
    if not text:
        return {}
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        return {}
    try:
        parsed = json.loads(text[start : end + 1])
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _flatten_bt_nodes(root: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}

    def walk(node: Dict[str, Any]) -> None:
        node_type = str(node.get("type", "") or "").strip().lower()
        node_id = str(node.get("id", "") or "").strip()
        if node_id and node_type in {"action", "condition"}:
            details = node.get(node_type, {}) if isinstance(node.get(node_type), dict) else {}
            out[node_id] = {
                "node_type": node_type,
                "description": str(details.get("description") or node.get("description") or ""),
                "success_criteria": str(details.get("success_criteria") or ""),
            }
        for child in node.get("children", []) or []:
            if isinstance(child, dict):
                walk(child)

    walk(root)
    return out


def _frame_token_from_paths(frame_paths: Any) -> str:
    if isinstance(frame_paths, list) and frame_paths:
        stem = Path(str(frame_paths[-1])).stem
        if stem.startswith("frame_"):
            return stem
    return ""


def _frame_token_from_entry(entry: Dict[str, Any]) -> str:
    token = str(entry.get("current_frame", "") or "")
    if token.startswith("frame_"):
        return token
    token = _frame_token_from_paths(entry.get("frames", []))
    if token:
        return token
    step = entry.get("step")
    if isinstance(step, int):
        return f"frame_{step:05d}"
    return ""


def _fmt_mmss_from_step(step: Optional[int], outer_step_seconds: float) -> str:
    if step is None:
        return ""
    total_seconds = max(0.0, float(step) * float(outer_step_seconds))
    minutes = int(total_seconds // 60)
    seconds = int(round(total_seconds - minutes * 60))
    if seconds == 60:
        minutes += 1
        seconds = 0
    return f"{minutes}:{seconds:02d}"


def _fmt_mmss_from_seconds(seconds: Optional[float]) -> str:
    if seconds is None:
        return ""
    total_seconds = max(0.0, float(seconds))
    minutes = int(total_seconds // 60)
    rem_seconds = int(round(total_seconds - (minutes * 60)))
    if rem_seconds == 60:
        minutes += 1
        rem_seconds = 0
    return f"{minutes}:{rem_seconds:02d}"


def _safe_video_duration_seconds(video_path: Path) -> Optional[float]:
    if not video_path.exists():
        return None

    try:
        import cv2  # type: ignore

        cap = cv2.VideoCapture(str(video_path))
        if cap.isOpened():
            frame_count = float(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0.0)
            fps = float(cap.get(cv2.CAP_PROP_FPS) or 0.0)
            cap.release()
            if fps > 0.0 and frame_count >= 0.0:
                return frame_count / fps
        else:
            cap.release()
    except Exception:
        pass

    try:
        import av  # type: ignore

        with av.open(str(video_path)) as container:
            if container.duration is not None:
                return float(container.duration) / 1_000_000.0
    except Exception:
        pass

    return None


def _find_primary_video_path(episode_dir: Path) -> Optional[Path]:
    video_dir = episode_dir / "video"
    if not video_dir.exists():
        return None
    mp4s = sorted(video_dir.glob("**/*.mp4"))
    return mp4s[0] if mp4s else None


def _parse_step_from_frame_token(token: Any) -> Optional[int]:
    text = str(token or "").strip()
    if not text.startswith("frame_"):
        return None
    try:
        return int(text.split("_", 1)[1])
    except Exception:
        return None


def _parse_mmss_to_seconds(text: Any) -> Optional[float]:
    raw = str(text or "").strip()
    if not raw or ":" not in raw:
        return None
    parts = raw.split(":")
    if len(parts) != 2:
        return None
    try:
        minutes = int(parts[0])
        seconds = int(parts[1])
    except Exception:
        return None
    if minutes < 0 or seconds < 0:
        return None
    return float(minutes * 60 + seconds)


def _backfill_sheet2_video_times(
    workbook_path: Path,
    outer_step_seconds: float,
    video_time_scale: Optional[float],
) -> int:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Title is more than 31 characters.*",
            category=UserWarning,
        )
        wb = load_workbook(workbook_path)

    if "Sheet2_MissionEngine_Review" not in wb.sheetnames:
        return 0
    ws = wb["Sheet2_MissionEngine_Review"]

    filled = 0
    start_row = 9
    for r in range(start_row, ws.max_row + 1):
        current_video_time = str(ws.cell(row=r, column=9).value or "").strip()
        if current_video_time:
            continue

        step = _parse_step_from_frame_token(ws.cell(row=r, column=7).value)
        machine_seconds: Optional[float]
        if step is not None:
            machine_seconds = float(step) * float(outer_step_seconds)
        else:
            machine_seconds = _parse_mmss_to_seconds(ws.cell(row=r, column=8).value)

        if machine_seconds is None:
            continue

        if video_time_scale is None:
            video_seconds = machine_seconds
        else:
            video_seconds = machine_seconds * float(video_time_scale)

        ws.cell(row=r, column=9, value=_fmt_mmss_from_seconds(video_seconds))
        filled += 1

    if filled > 0:
        wb.save(workbook_path)
    return filled


def _derive_timing_from_summary(
    summary: Dict[str, Any],
    status_eval_seconds_override: Optional[float],
    status_window_seconds_override: Optional[float],
    status_window_frames_override: Optional[int],
) -> Tuple[int, int]:
    outer_step_seconds = float(summary.get("outer_step_seconds", 0.4) or 0.4)
    default_eval_seconds = float(summary.get("status_eval_seconds_target", 2.0) or 2.0)
    default_window_seconds = float(summary.get("status_window_seconds_target", 4.0) or 4.0)
    default_window_frames = int(summary.get("status_window_frames", 0) or 0)

    eval_seconds = float(status_eval_seconds_override) if status_eval_seconds_override is not None else default_eval_seconds
    window_seconds = (
        float(status_window_seconds_override)
        if status_window_seconds_override is not None
        else default_window_seconds
    )

    n_action_steps = max(1, int(round(outer_step_seconds * 20.0)))
    cfg = PipelineConfig(
        env_name=str(summary.get("env_name", "replay")),
        n_action_steps=n_action_steps,
        control_freq_hz=20.0,
        status_eval_seconds=eval_seconds,
        status_window_seconds=window_seconds,
        status_window_frames=default_window_frames,
    )

    eval_every_steps = int(summary.get("status_eval_every_n_steps", 0) or 0)
    if eval_every_steps <= 0:
        eval_every_steps = _status_eval_every_steps(cfg)

    if status_window_frames_override is not None and int(status_window_frames_override) > 0:
        window_frames = int(status_window_frames_override)
    else:
        window_frames = int(summary.get("status_window_frames", 0) or 0)
        if window_frames <= 0:
            window_frames = _status_window_frame_count(cfg)

    return max(1, eval_every_steps), max(1, window_frames)


def _frame_index_from_path(path: Path) -> Optional[int]:
    stem = path.stem
    if not stem.startswith("frame_"):
        return None
    try:
        return int(stem.split("_", 1)[1])
    except Exception:
        return None


def _build_manifest_rows_from_frames(
    episode_dir: Path,
    episode_id: str,
    eval_every_steps: int,
    window_frames: int,
    final_step: Optional[int],
) -> List[Dict[str, Any]]:
    frame_dir = episode_dir / "frames"
    if not frame_dir.exists():
        return []

    indexed_paths: List[Tuple[int, Path]] = []
    for p in sorted(frame_dir.glob("frame_*.jpg")):
        idx = _frame_index_from_path(p)
        if idx is not None:
            indexed_paths.append((idx, p))
    if not indexed_paths:
        return []

    max_step_from_frames = indexed_paths[-1][0]
    step_limit = max_step_from_frames
    if final_step is not None:
        step_limit = min(step_limit, int(final_step))
    if step_limit < 0:
        return []

    path_by_step = {idx: p for idx, p in indexed_paths}
    rows: List[Dict[str, Any]] = []

    step = max(1, int(eval_every_steps))
    while step <= step_limit:
        start = max(0, step - int(window_frames) + 1)
        window_paths: List[str] = []
        for s in range(start, step + 1):
            p = path_by_step.get(s)
            if p is not None:
                window_paths.append(str(p))
        if window_paths:
            rows.append(
                {
                    "ts": 0.0,
                    "schema": "dabtroll.status_window_manifest.v1",
                    "status_eval_id": f"{episode_id}:step_{step}:replay_from_frames",
                    "episode_id": episode_id,
                    "step": int(step),
                    "frames": window_paths,
                    "prompt_text": "",
                }
            )
        step += max(1, int(eval_every_steps))

    if step_limit not in {row.get("step") for row in rows}:
        start = max(0, int(step_limit) - int(window_frames) + 1)
        window_paths = [str(path_by_step[s]) for s in range(start, int(step_limit) + 1) if s in path_by_step]
        if window_paths:
            rows.append(
                {
                    "ts": 0.0,
                    "schema": "dabtroll.status_window_manifest.v1",
                    "status_eval_id": f"{episode_id}:step_{int(step_limit)}:replay_terminal",
                    "episode_id": episode_id,
                    "step": int(step_limit),
                    "frames": window_paths,
                    "prompt_text": "",
                }
            )

    rows.sort(key=lambda r: int(r.get("step", 10**9)))
    return rows


def _filter_manifest_for_simulation_timing(
    manifest_rows: List[Dict[str, Any]],
    expected_eval_every_steps: int,
    expected_window_frames: int,
    final_step: Optional[int],
) -> List[Dict[str, Any]]:
    filtered: List[Dict[str, Any]] = []
    first_step: Optional[int] = None
    for row in manifest_rows:
        s = row.get("step")
        if isinstance(s, int):
            first_step = s
            break

    for row in manifest_rows:
        step = row.get("step")
        if not isinstance(step, int):
            continue

        frame_count = len(row.get("frames", [])) if isinstance(row.get("frames"), list) else 0
        if frame_count < expected_window_frames:
            continue

        if first_step is None:
            at_expected_cadence = True
        else:
            at_expected_cadence = ((step - first_step) % max(1, expected_eval_every_steps)) == 0
        is_terminal_eval = final_step is not None and step >= final_step
        if at_expected_cadence or is_terminal_eval:
            filtered.append(row)

    if filtered:
        return filtered
    return manifest_rows


def _style_header_row(ws, row_idx: int, columns: Iterable[int]) -> None:
    for col in columns:
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _style_data_grid(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c in (12, 18):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def _build_mission_milestones(
    mission_rows: List[Dict[str, Any]],
    btstatus_rows: List[Dict[str, Any]],
    outer_step_seconds: float,
    video_time_scale: Optional[float],
) -> List[Dict[str, Any]]:
    milestones: List[Dict[str, Any]] = []

    def _video_time_from_step(step: Optional[int]) -> str:
        if step is None:
            return ""
        machine_seconds = max(0.0, float(step) * float(outer_step_seconds))
        if video_time_scale is None:
            return _fmt_mmss_from_seconds(machine_seconds)
        return _fmt_mmss_from_seconds(machine_seconds * float(video_time_scale))

    for row in mission_rows:
        if str(row.get("direction", "") or "") != "response":
            continue
        request_kind = str(row.get("request_kind", "") or "")
        if "status_eval" not in request_kind:
            continue

        response = row.get("response") if isinstance(row.get("response"), dict) else {}
        response_text = str(row.get("response_text", "") or response.get("text", "") or "")
        parsed = _parse_status_json(response_text)
        step = row.get("step") if isinstance(row.get("step"), int) else None

        model_status = str(parsed.get("status", "") or "").strip().lower()
        model_progress = str(parsed.get("progress_score", parsed.get("progress", "")) or "")
        model_notes = str(parsed.get("notes", "") or "")
        if not model_status:
            response_ok = bool(row.get("response_ok", response.get("ok", False)))
            response_error = str(row.get("response_error", response.get("error", "")) or "").strip()
            if response_ok and response_text.strip():
                model_status = "unparsed"
                if not model_notes:
                    model_notes = response_text.strip()[:500]
            else:
                model_status = "error"
                if response_error:
                    model_notes = response_error
                elif response_text.strip():
                    model_notes = response_text.strip()[:500]
                else:
                    model_notes = "empty response"

        milestones.append(
            {
                "sort_step": step if step is not None else 10**9,
                "sort_ts": float(row.get("ts", 0.0)) if isinstance(row.get("ts"), (int, float)) else 0.0,
                "source": "mission_engine_response",
                "request_kind": request_kind,
                "status_eval_id": str(row.get("status_eval_id", "") or ""),
                "node_id": str(row.get("node_id", "") or ""),
                "node_type": str(row.get("node_type", "") or ""),
                "frame": _frame_token_from_entry(row),
                "time": _fmt_mmss_from_step(step, outer_step_seconds),
                "video_time": _video_time_from_step(step),
                "model_status": model_status,
                "model_progress": model_progress,
                "model_notes": model_notes,
                "model_latency_s": response.get("latency_s", row.get("response_latency_s", "")),
            }
        )

    for row in btstatus_rows:
        status = str(row.get("status", "") or "").strip().lower()
        if status != "complete":
            continue
        step = row.get("step") if isinstance(row.get("step"), int) else None
        milestones.append(
            {
                "sort_step": step if step is not None else 10**9,
                "sort_ts": float(row.get("ts", 0.0)) if isinstance(row.get("ts"), (int, float)) else 0.0,
                "source": "btstatus_complete_milestone",
                "request_kind": "node_completion",
                "status_eval_id": str(row.get("status_eval_id", "") or ""),
                "node_id": str(row.get("node_id", "") or ""),
                "node_type": str(row.get("node_type", "") or ""),
                "frame": _frame_token_from_paths(row.get("frames", [])),
                "time": _fmt_mmss_from_step(step, outer_step_seconds),
                "video_time": _video_time_from_step(step),
                "model_status": str(row.get("status", "") or ""),
                "model_progress": str(row.get("progress_score", "") or ""),
                "model_notes": str(row.get("notes", "") or ""),
                "model_latency_s": "",
            }
        )

    milestones.sort(key=lambda m: (m["sort_step"], m["sort_ts"], m["source"]))
    return milestones


def _write_sheet3(workbook_path: Path, summary: Dict[str, Any], milestones: List[Dict[str, Any]], bt_svg_path: Path) -> None:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Title is more than 31 characters.*",
            category=UserWarning,
        )
        wb = load_workbook(workbook_path)
    sheet_name = "sheet3_qwen3_5_MissionEngine_Review"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Title is more than 31 characters.*",
            category=UserWarning,
        )
        ws = wb.create_sheet(sheet_name)

    ws.merge_cells("A1:S1")
    ws["A1"] = "Human Review Sheet 3: Qwen3.5 Mission Engine Response and Milestone Agreement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = TITLE_FILL

    instructions = (
        "Instructions: Review major milestones (especially node completion points). Video time starts at 0:00 for "
        "frame_00000. For each row, judge whether qwen3.5 mission-engine assessment is correct, enter reviewer time/frame "
        "if different, and explain disagreement."
    )
    ws.merge_cells("A2:S4")
    ws["A2"] = instructions
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A5:S5")
    ws["A5"] = f"BT SVG (open externally): {bt_svg_path}"
    ws["A5"].hyperlink = bt_svg_path.as_uri() if bt_svg_path.exists() else None
    ws["A5"].font = Font(color="0563C1", underline="single") if bt_svg_path.exists() else Font(bold=True)

    ws.merge_cells("A6:S6")
    ws["A6"] = (
        f"Episode: {summary.get('episode_id', '')} | Mission: {summary.get('mission_name', '')} | "
        f"Model: Qwen/Qwen3.5-9B"
    )

    headers = [
        "milestone_idx",
        "source",
        "request_kind_or_event",
        "status_eval_id",
        "node_id",
        "node_type",
        "machine_frame",
        "machine_time_m:ss",
        "video_time_m:ss",
        "model_status",
        "model_progress",
        "model_notes",
        "model_latency_s",
        "reviewer_agree_YN_partial",
        "reviewer_time_m:ss",
        "reviewer_frame",
        "reviewer_status",
        "reviewer_notes",
        "evidence_refs",
    ]

    header_row = 8
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=h)
    _style_header_row(ws, header_row, range(1, len(headers) + 1))

    row_idx = header_row + 1
    for i, m in enumerate(milestones, start=1):
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=m.get("source", ""))
        ws.cell(row=row_idx, column=3, value=m.get("request_kind", ""))
        ws.cell(row=row_idx, column=4, value=m.get("status_eval_id", ""))
        ws.cell(row=row_idx, column=5, value=m.get("node_id", ""))
        ws.cell(row=row_idx, column=6, value=m.get("node_type", ""))
        ws.cell(row=row_idx, column=7, value=m.get("frame", ""))
        ws.cell(row=row_idx, column=8, value=m.get("time", ""))
        ws.cell(row=row_idx, column=9, value=m.get("video_time", ""))
        ws.cell(row=row_idx, column=10, value=m.get("model_status", ""))
        ws.cell(row=row_idx, column=11, value=m.get("model_progress", ""))
        ws.cell(row=row_idx, column=12, value=m.get("model_notes", ""))
        ws.cell(row=row_idx, column=13, value=m.get("model_latency_s", ""))
        row_idx += 1

    data_end = max(header_row + 1, row_idx - 1)
    _style_data_grid(ws, header_row, data_end, 1, len(headers))

    widths = {
        1: 12,
        2: 26,
        3: 30,
        4: 34,
        5: 28,
        6: 12,
        7: 16,
        8: 18,
        9: 16,
        10: 16,
        11: 14,
        12: 42,
        13: 14,
        14: 24,
        15: 20,
        16: 18,
        17: 18,
        18: 40,
        19: 34,
    }
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    ws.freeze_panes = "A9"
    wb.save(workbook_path)


def _replay_episode(
    episode_dir: Path,
    mission_host: str,
    mission_port: int,
    mission_timeout_ms: int,
    out_tag: str,
    status_eval_seconds_override: Optional[float],
    status_window_seconds_override: Optional[float],
    status_window_frames_override: Optional[int],
    use_manifest_prompts: bool,
    window_source: str,
) -> Dict[str, Any]:
    summary_path = episode_dir / "episode_summary.json"
    bt_path = episode_dir / "bt.json"
    manifest_path = episode_dir / "status_window_manifest.jsonl"
    workbook_path = episode_dir / "human_rater_evaluation.xlsx"

    if not summary_path.exists() or not bt_path.exists() or not manifest_path.exists() or not workbook_path.exists():
        raise FileNotFoundError(
            f"Episode missing required files in {episode_dir}: episode_summary.json, bt.json, status_window_manifest.jsonl, human_rater_evaluation.xlsx"
        )

    summary = _read_json(summary_path)
    bt_json = _read_json(bt_path)
    bt_root = bt_json.get("root", {}) if isinstance(bt_json, dict) else {}
    node_map = _flatten_bt_nodes(bt_root if isinstance(bt_root, dict) else {})
    runner = BehaviorTreeRunner(bt_json if isinstance(bt_json, dict) else {"root": {}}, mission_name=str(summary.get("mission_name", "")))
    runner.reset()

    input_manifest_rows = _read_jsonl(manifest_path)
    input_manifest_rows.sort(key=lambda r: (int(r.get("step", 10**9)) if isinstance(r.get("step"), int) else 10**9, float(r.get("ts", 0.0)) if isinstance(r.get("ts"), (int, float)) else 0.0))

    expected_eval_every_steps, expected_window_frames = _derive_timing_from_summary(
        summary=summary,
        status_eval_seconds_override=status_eval_seconds_override,
        status_window_seconds_override=status_window_seconds_override,
        status_window_frames_override=status_window_frames_override,
    )
    final_step = int(summary.get("steps_executed", 0) or 0) - 1
    if str(window_source).strip().lower() == "frames":
        generated_rows = _build_manifest_rows_from_frames(
            episode_dir=episode_dir,
            episode_id=str(summary.get("episode_id", "") or ""),
            eval_every_steps=expected_eval_every_steps,
            window_frames=expected_window_frames,
            final_step=final_step if final_step >= 0 else None,
        )
        if generated_rows:
            input_manifest_rows = generated_rows
    else:
        input_manifest_rows = _filter_manifest_for_simulation_timing(
            manifest_rows=input_manifest_rows,
            expected_eval_every_steps=expected_eval_every_steps,
            expected_window_frames=expected_window_frames,
            final_step=final_step if final_step >= 0 else None,
        )

    qwen_out_dir = episode_dir / f"qwen_3_5_{out_tag}"
    qwen_out_dir.mkdir(parents=True, exist_ok=True)
    out_manifest_path = qwen_out_dir / "status_window_manifest.jsonl"
    out_mission_path = qwen_out_dir / "missionengine.jsonl"
    out_trace_path = qwen_out_dir / "pipeline_trace.jsonl"

    client = MissionEngineClient(host=mission_host, port=mission_port, timeout_ms=mission_timeout_ms)

    replay_manifest_rows: List[Dict[str, Any]] = []
    replay_mission_rows: List[Dict[str, Any]] = []
    replay_trace_rows: List[Dict[str, Any]] = []
    btstatus_complete_rows: List[Dict[str, Any]] = []

    for item in input_manifest_rows:
        bt_state, active_node = runner.tick()
        if bt_state in {"complete", "failure"} or not isinstance(active_node, dict):
            break

        step = int(item.get("step", 0) or 0)
        node_id = str(active_node.get("id", "") or "")
        node_type = str(active_node.get("type", "") or "")
        status_eval_id = str(item.get("status_eval_id", "") or f"qwen35:{summary.get('episode_id','')}:step_{step}:{node_id}")
        frames = [str(p) for p in item.get("frames", []) if str(p)]
        frame_paths = [Path(p) for p in frames if Path(p).exists()]
        if not frame_paths:
            continue

        prompt_text = str(item.get("prompt_text", "") or "").strip() if use_manifest_prompts else ""
        if not prompt_text:
            node_info = node_map.get(node_id, {})
            prompt_text = status_eval_text(
                node_type or str(node_info.get("node_type", "") or "action"),
                str(node_info.get("description", "") or ""),
                str(node_info.get("success_criteria", "") or ""),
            )

        frames_b64 = [_encode_jpg_file(p) for p in frame_paths]
        current_frame = _frame_token_from_paths(frames)
        current_frame_path = str(frame_paths[-1])

        request_payload = {
            "mode": "text_video",
            "prompt_text": prompt_text,
            "frames_b64": frames_b64,
        }

        req_ts = time.time()
        req_row = {
            "ts": req_ts,
            "direction": "request",
            "request_kind": "status_eval_qwen3_5",
            "episode_id": summary.get("episode_id", ""),
            "step": step,
            "current_frame": current_frame,
            "current_frame_path": current_frame_path,
            "frames": frames,
            "frames_count": len(frames),
            "node_id": node_id,
            "node_type": node_type,
            "status_eval_id": status_eval_id,
            "request": {"mode": "text_video", "prompt_text": prompt_text},
        }
        replay_mission_rows.append(req_row)

        response = client.request(request_payload, retries=0)
        resp_ts = time.time()
        response_text = str(response.get("text", "") or "") if isinstance(response, dict) else ""
        parsed = _parse_status_json(response_text)

        resp_row = {
            "ts": resp_ts,
            "direction": "response",
            "request_kind": "status_eval_qwen3_5",
            "episode_id": summary.get("episode_id", ""),
            "step": step,
            "current_frame": current_frame,
            "current_frame_path": current_frame_path,
            "frames": frames,
            "frames_count": len(frames),
            "node_id": node_id,
            "node_type": node_type,
            "status_eval_id": status_eval_id,
            "response": response if isinstance(response, dict) else {"ok": False, "error": "non_dict_response"},
            "response_ok": bool(response.get("ok")) if isinstance(response, dict) else False,
            "response_latency_s": response.get("latency_s", "") if isinstance(response, dict) else "",
            "response_error": response.get("error") if isinstance(response, dict) else "non_dict_response",
            "response_text": response_text,
        }
        replay_mission_rows.append(resp_row)

        out_manifest = dict(item)
        out_manifest["qwen35_replay"] = True
        out_manifest["qwen35_ts"] = resp_ts
        out_manifest["qwen35_status_eval_id"] = status_eval_id
        out_manifest["qwen35_response_text"] = response_text
        out_manifest["qwen35_response_ok"] = bool(response.get("ok")) if isinstance(response, dict) else False
        out_manifest["qwen35_status"] = str(parsed.get("status", "") or "")
        out_manifest["qwen35_progress_score"] = parsed.get("progress_score", parsed.get("progress", ""))
        out_manifest["qwen35_notes"] = str(parsed.get("notes", "") or "")
        replay_manifest_rows.append(out_manifest)

        status = str(parsed.get("status", "") or "")
        progress = parsed.get("progress_score", parsed.get("progress", ""))
        notes = str(parsed.get("notes", "") or "")
        trace_event = {
            "ts": resp_ts,
            "event": "status_eval_qwen3_5",
            "episode_id": summary.get("episode_id", ""),
            "status_eval_id": status_eval_id,
            "step": step,
            "node_id": node_id,
            "node_type": node_type,
            "status": status,
            "status_label_std": status.strip().lower(),
            "vlm_confidence": 1.0,
            "notes": notes,
            "progress_score": progress,
            "criteria_met": parsed.get("criteria_met", []) if isinstance(parsed.get("criteria_met", []), list) else [],
            "criteria_missing": parsed.get("criteria_missing", []) if isinstance(parsed.get("criteria_missing", []), list) else [],
            "qwen_model": "Qwen/Qwen3.5-9B",
        }
        replay_trace_rows.append(trace_event)

        if status.strip().lower() == "complete":
            btstatus_complete_rows.append(
                {
                    "ts": resp_ts,
                    "step": step,
                    "node_id": node_id,
                    "node_type": node_type,
                    "status": "complete",
                    "progress_score": progress,
                    "notes": notes,
                    "status_eval_id": status_eval_id,
                    "frames": frames,
                }
            )

        if status.strip().lower() in {"complete", "failure"}:
            runner.set_status(active_node, {"status": status.strip().lower()})

    _write_jsonl(out_manifest_path, replay_manifest_rows)
    _write_jsonl(out_mission_path, replay_mission_rows)
    _write_jsonl(out_trace_path, replay_trace_rows)

    outer_step_seconds = float(summary.get("outer_step_seconds", 0.4) or 0.4)
    steps_executed = int(summary.get("steps_executed", 0) or 0)
    machine_duration_seconds = max(0.0, float(steps_executed) * outer_step_seconds)
    video_duration_seconds: Optional[float] = None
    video_time_scale: Optional[float] = None
    video_path = _find_primary_video_path(episode_dir)
    if video_path is not None:
        video_duration_seconds = _safe_video_duration_seconds(video_path)
    if video_duration_seconds is not None and machine_duration_seconds > 0.0:
        video_time_scale = float(video_duration_seconds) / float(machine_duration_seconds)

    milestones = _build_mission_milestones(
        mission_rows=replay_mission_rows,
        btstatus_rows=btstatus_complete_rows,
        outer_step_seconds=outer_step_seconds,
        video_time_scale=video_time_scale,
    )
    _write_sheet3(
        workbook_path=workbook_path,
        summary=summary,
        milestones=milestones,
        bt_svg_path=episode_dir / "bt.svg",
    )

    sheet2_video_times_filled = _backfill_sheet2_video_times(
        workbook_path=workbook_path,
        outer_step_seconds=outer_step_seconds,
        video_time_scale=video_time_scale,
    )

    return {
        "episode_dir": str(episode_dir),
        "qwen_output_dir": str(qwen_out_dir),
        "window_source": str(window_source),
        "expected_eval_every_n_steps": int(expected_eval_every_steps),
        "expected_status_window_frames": int(expected_window_frames),
        "manifest_rows": len(replay_manifest_rows),
        "mission_rows": len(replay_mission_rows),
        "trace_rows": len(replay_trace_rows),
        "sheet2_video_times_filled": int(sheet2_video_times_filled),
        "sheet3_added": str(workbook_path),
    }


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Replay Qwen3.5 mission evaluation over existing episode frames")
    ap.add_argument("--episode-dir", action="append", default=[], help="Episode directory (repeatable)")
    ap.add_argument("--mission-host", default="127.0.0.1")
    ap.add_argument("--mission-port", type=int, default=5560)
    ap.add_argument("--mission-timeout-ms", type=int, default=120000)
    ap.add_argument(
        "--status-eval-seconds",
        type=float,
        default=None,
        help="Optional override for eval cadence seconds. Default: use episode_summary status_eval_seconds_target.",
    )
    ap.add_argument(
        "--status-window-seconds",
        type=float,
        default=None,
        help="Optional override for status window seconds. Default: use episode_summary status_window_seconds_target.",
    )
    ap.add_argument(
        "--status-window-frames",
        type=int,
        default=None,
        help="Optional override for exact window frame count.",
    )
    ap.add_argument(
        "--ignore-manifest-prompts",
        action="store_true",
        help="Rebuild prompts from BT metadata instead of reusing prompt_text from status_window_manifest.",
    )
    ap.add_argument(
        "--window-source",
        choices=["frames", "manifest"],
        default="frames",
        help="Use full frames directory (frames, default) or status_window_manifest timing rows (manifest).",
    )
    return ap.parse_args()


def main() -> None:
    args = _parse_args()
    if not args.episode_dir:
        raise SystemExit("Provide at least one --episode-dir")

    out_tag = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    results: List[Dict[str, Any]] = []
    for ep in args.episode_dir:
        ep_path = Path(ep).expanduser().resolve()
        results.append(
            _replay_episode(
                episode_dir=ep_path,
                mission_host=args.mission_host,
                mission_port=int(args.mission_port),
                mission_timeout_ms=int(args.mission_timeout_ms),
                out_tag=out_tag,
                status_eval_seconds_override=args.status_eval_seconds,
                status_window_seconds_override=args.status_window_seconds,
                status_window_frames_override=args.status_window_frames,
                use_manifest_prompts=not bool(args.ignore_manifest_prompts),
                window_source=str(args.window_source),
            )
        )

    print(json.dumps({"ok": True, "timestamp": out_tag, "episodes": results}, indent=2))


if __name__ == "__main__":
    main()
