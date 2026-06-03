from __future__ import annotations

"""Generate a balanced human-eval schedule workbook.

The workbook is designed for 3 scenarios with 20 eval rows each (60 total).
It samples episodes 2..21 only, balances picks across runs, and can preserve
already-completed rows from an existing workbook at the top.
"""

import argparse
import datetime as dt
import json
import random
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ModuleNotFoundError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: python -m pip install openpyxl"
    ) from exc


HEADER_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _norm_col(name: object) -> str:
    return "".join(ch if str(ch).isalnum() else "_" for ch in str(name).strip().lower()).strip("_")


def _is_nonempty_signal(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text not in {"", "nan", "none"}


def _find_column_index(headers: Sequence[object], aliases: Sequence[str]) -> Optional[int]:
    norm_to_idx: Dict[str, int] = {}
    for idx, header in enumerate(headers, start=1):
        norm_to_idx[_norm_col(header)] = idx

    for alias in aliases:
        alias_norm = _norm_col(alias)
        if alias_norm in norm_to_idx:
            return norm_to_idx[alias_norm]

    for alias in aliases:
        alias_norm = _norm_col(alias)
        for norm_key, idx in norm_to_idx.items():
            if alias_norm in norm_key:
                return idx
    return None


def _find_header_row_for_signals(ws) -> Optional[int]:
    for row_idx in (1, 8, 9):
        if row_idx > ws.max_row:
            continue
        headers = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        norm_headers = [_norm_col(h) for h in headers]
        for token in ("reviewer", "model_status", "node_id", "status_eval_id"):
            if any(token in h for h in norm_headers):
                return row_idx
    return None


def _sheet_has_reviewer_signal(ws, aliases: Sequence[Sequence[str]]) -> bool:
    header_row = _find_header_row_for_signals(ws)
    if header_row is None:
        return False

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    candidate_cols: List[int] = []
    for alias_group in aliases:
        col_idx = _find_column_index(headers, alias_group)
        if col_idx is not None:
            candidate_cols.append(col_idx)

    if not candidate_cols:
        return False

    for col_idx in candidate_cols:
        for r in range(header_row + 1, ws.max_row + 1):
            if _is_nonempty_signal(ws.cell(row=r, column=col_idx).value):
                return True
    return False


def _episode_human_eval_complete(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return False

    sheets = wb.worksheets
    if not sheets:
        return False

    s1_aliases = [
        ["reviewer_completion_time_m:ss", "reviewer_completion_time", "reviewer_time_m:ss"],
        ["reviewer_agrees_machine_yn_partial", "reviewer_agree_yn_partial"],
        ["reviewer_confidence_1to5", "reviewer_confidence"],
        ["reviewer_notes"],
    ]
    s2_aliases = [
        ["reviewer_agree_yn_partial", "reviewer_agrees_machine_yn_partial"],
        ["reviewer_time_m:ss", "reviewer_completion_time_m:ss"],
        ["reviewer_status"],
        ["reviewer_progress"],
        ["reviewer_notes"],
    ]

    if _sheet_has_reviewer_signal(sheets[0], s1_aliases):
        return True
    if len(sheets) > 1 and _sheet_has_reviewer_signal(sheets[1], s2_aliases):
        return True
    return False


@dataclass(frozen=True)
class EpisodeRecord:
    scenario_id: str
    run_tag: str
    episode_index: int
    mission_name: str
    episode_id: str
    episode_dir: Path
    video_path: Optional[Path]
    survey_path: Optional[Path]

    @property
    def key(self) -> Tuple[str, str, int]:
        return (self.scenario_id, self.run_tag, self.episode_index)


@dataclass
class ScheduleRow:
    scenario_id: str
    run_tag: str
    episode_index: int
    mission_name: str
    episode_id: str
    episode_dir: Path
    video_path: Optional[Path]
    survey_path: Optional[Path]
    complete: bool = False
    source: str = "new"

    @property
    def key(self) -> Tuple[str, str, int]:
        return (self.scenario_id, self.run_tag, self.episode_index)


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Generate balanced human-eval schedule workbook")
    ap.add_argument("--project-root", default="/home/mark/dabtroll", help="Project root path")
    ap.add_argument("--logs-dir", default="", help="Logs directory (default: <project-root>/data/logs)")
    ap.add_argument(
        "--scenarios",
        nargs="*",
        default=[],
        help="Exactly 3 scenario ids to include. If omitted, auto-detect top 3 by available runs.",
    )
    ap.add_argument("--evals-per-scenario", type=int, default=20, help="Rows per scenario")
    ap.add_argument(
        "--runs-per-scenario",
        type=int,
        default=5,
        help="Number of runs to sample per scenario (default: 5)",
    )
    ap.add_argument("--seed", type=int, default=7, help="Random seed for reproducibility")
    ap.add_argument(
        "--existing-workbook",
        default="",
        help="Optional existing workbook path. Completed rows are moved to the top in the new workbook.",
    )
    ap.add_argument(
        "--output",
        default="",
        help="Output xlsx path (default: <logs-dir>/human_eval_schedule_<UTC timestamp>.xlsx)",
    )
    return ap.parse_args()


def _find_video_path(episode_dir: Path) -> Optional[Path]:
    video_dir = episode_dir / "video"
    if not video_dir.exists():
        return None
    mp4s = sorted(video_dir.glob("**/*.mp4"))
    return mp4s[0] if mp4s else None


def _find_survey_path(episode_dir: Path) -> Optional[Path]:
    candidate = episode_dir / "human_rater_evaluation.xlsx"
    return candidate if candidate.exists() else None


def _load_summary_records(logs_dir: Path) -> List[EpisodeRecord]:
    records: List[EpisodeRecord] = []
    for summary_path in sorted(logs_dir.glob("simulation_summary_dabtroll_*.json")):
        try:
            payload = json.loads(summary_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(payload, list):
            continue
        for item in payload:
            if not isinstance(item, dict):
                continue
            scenario_id = str(item.get("scenario_id") or "").strip()
            run_tag = str(item.get("run_tag") or "").strip()
            episode_index = item.get("episode_index")
            mission_name = str(item.get("mission_name") or "").strip()
            episode_id = str(item.get("episode_id") or "").strip()
            episode_dir_raw = str(item.get("episode_dir") or "").strip()
            if not scenario_id or not run_tag or not episode_dir_raw:
                continue
            if not isinstance(episode_index, int):
                continue
            if episode_index < 2 or episode_index > 21:
                continue
            episode_dir = Path(episode_dir_raw)
            rec = EpisodeRecord(
                scenario_id=scenario_id,
                run_tag=run_tag,
                episode_index=episode_index,
                mission_name=mission_name,
                episode_id=episode_id,
                episode_dir=episode_dir,
                video_path=_find_video_path(episode_dir),
                survey_path=_find_survey_path(episode_dir),
            )
            records.append(rec)

    dedup: Dict[Tuple[str, str, int], EpisodeRecord] = {}
    for rec in records:
        dedup[rec.key] = rec
    return list(dedup.values())


def _group_by_scenario_run(records: Sequence[EpisodeRecord]) -> Dict[str, Dict[str, List[EpisodeRecord]]]:
    grouped: Dict[str, Dict[str, List[EpisodeRecord]]] = {}
    for rec in records:
        grouped.setdefault(rec.scenario_id, {}).setdefault(rec.run_tag, []).append(rec)
    for run_map in grouped.values():
        for run_tag in list(run_map.keys()):
            run_map[run_tag] = sorted(run_map[run_tag], key=lambda r: r.episode_index)
    return grouped


def _choose_scenarios(
    grouped: Dict[str, Dict[str, List[EpisodeRecord]]],
    requested: Sequence[str],
) -> List[str]:
    if requested:
        if len(requested) != 3:
            raise ValueError("You must pass exactly 3 scenario ids with --scenarios.")
        missing = [s for s in requested if s not in grouped]
        if missing:
            raise ValueError(f"Requested scenario(s) missing in summaries: {missing}")
        return list(requested)

    ranked = sorted(
        grouped.items(),
        key=lambda kv: (len(kv[1]), sum(len(v) for v in kv[1].values())),
        reverse=True,
    )
    if len(ranked) < 3:
        raise ValueError("Need at least 3 scenarios in logs to auto-select.")
    return [scenario for scenario, _ in ranked[:3]]


def _normalize_complete(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "done", "complete", "completed", "x"}


def _load_completed_rows(
    existing_workbook: Optional[Path],
    selected_scenarios: Sequence[str],
    valid_keys: Iterable[Tuple[str, str, int]],
) -> List[ScheduleRow]:
    if existing_workbook is None or not existing_workbook.exists():
        return []

    valid_key_set = set(valid_keys)
    wb = load_workbook(existing_workbook, data_only=True)
    ws = wb.active

    header_row = None
    headers: Dict[str, int] = {}
    expected = {"complete", "scenario_id", "run_tag", "episode_index"}

    for r in range(1, min(25, ws.max_row) + 1):
        found: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            key = str(val or "").strip().lower()
            if key in expected:
                found[key] = c
        if expected.issubset(found.keys()):
            header_row = r
            headers = found
            break

    if header_row is None:
        return []

    optional_cols: Dict[str, int] = {}
    known_optional = ["mission_name", "episode_id", "episode_dir", "video_link", "survey_link"]
    for c in range(1, ws.max_column + 1):
        key = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
        if key in known_optional:
            optional_cols[key] = c

    completed: List[ScheduleRow] = []
    allowed_scenarios = set(selected_scenarios)
    for r in range(header_row + 1, ws.max_row + 1):
        complete_val = ws.cell(row=r, column=headers["complete"]).value
        if not _normalize_complete(complete_val):
            continue

        scenario_id = str(ws.cell(row=r, column=headers["scenario_id"]).value or "").strip()
        run_tag = str(ws.cell(row=r, column=headers["run_tag"]).value or "").strip()
        episode_raw = ws.cell(row=r, column=headers["episode_index"]).value
        try:
            episode_index = int(episode_raw)
        except Exception:
            continue

        key = (scenario_id, run_tag, episode_index)
        if scenario_id not in allowed_scenarios or key not in valid_key_set:
            continue

        mission_name = (
            str(ws.cell(row=r, column=optional_cols["mission_name"]).value or "").strip()
            if "mission_name" in optional_cols
            else ""
        )
        episode_id = (
            str(ws.cell(row=r, column=optional_cols["episode_id"]).value or "").strip()
            if "episode_id" in optional_cols
            else ""
        )
        episode_dir_text = (
            str(ws.cell(row=r, column=optional_cols["episode_dir"]).value or "").strip()
            if "episode_dir" in optional_cols
            else ""
        )
        episode_dir = Path(episode_dir_text) if episode_dir_text else Path("")

        completed.append(
            ScheduleRow(
                scenario_id=scenario_id,
                run_tag=run_tag,
                episode_index=episode_index,
                mission_name=mission_name,
                episode_id=episode_id,
                episode_dir=episode_dir,
                video_path=None,
                survey_path=None,
                complete=True,
                source="existing_complete",
            )
        )

    return completed


def _load_episode_workbook_completions(
    records: Sequence[EpisodeRecord],
    selected_scenarios: Sequence[str],
) -> List[ScheduleRow]:
    allowed = set(selected_scenarios)
    completed: List[ScheduleRow] = []

    for rec in records:
        if rec.scenario_id not in allowed:
            continue
        if rec.survey_path is None:
            continue
        if not _episode_human_eval_complete(rec.survey_path):
            continue
        completed.append(
            ScheduleRow(
                scenario_id=rec.scenario_id,
                run_tag=rec.run_tag,
                episode_index=rec.episode_index,
                mission_name=rec.mission_name,
                episode_id=rec.episode_id,
                episode_dir=rec.episode_dir,
                video_path=rec.video_path,
                survey_path=rec.survey_path,
                complete=True,
                source="episode_workbook_complete",
            )
        )
    return completed


def _pick_rows_for_scenario(
    scenario_runs: Dict[str, List[EpisodeRecord]],
    total_needed: int,
    completed_keys: Iterable[Tuple[str, str, int]],
    rng: random.Random,
) -> List[EpisodeRecord]:
    completed = set(completed_keys)
    runs = sorted(scenario_runs.keys())
    if not runs:
        raise ValueError("Scenario has no runs available.")

    available_by_run: Dict[str, List[EpisodeRecord]] = {}
    for run_tag in runs:
        available_by_run[run_tag] = [r for r in scenario_runs[run_tag] if r.key not in completed]

    chosen: Dict[str, int] = {run_tag: 0 for run_tag in runs}

    while sum(chosen.values()) < total_needed:
        candidates = [
            run_tag
            for run_tag in runs
            if chosen[run_tag] < len(available_by_run[run_tag])
        ]
        if not candidates:
            raise ValueError(
                f"Not enough available episodes to satisfy scenario quota ({total_needed})."
            )

        min_count = min(chosen[run_tag] for run_tag in candidates)
        tied = [run_tag for run_tag in candidates if chosen[run_tag] == min_count]
        run_pick = rng.choice(tied)
        chosen[run_pick] += 1

    picked: List[EpisodeRecord] = []
    for run_tag in runs:
        need = chosen[run_tag]
        if need <= 0:
            continue
        pool = list(available_by_run[run_tag])
        rng.shuffle(pool)
        picked.extend(pool[:need])
    return picked


def _select_runs_for_scenario(
    scenario_runs: Dict[str, List[EpisodeRecord]],
    runs_per_scenario: int,
    required_run_tags: Optional[Sequence[str]] = None,
) -> Dict[str, List[EpisodeRecord]]:
    if runs_per_scenario <= 0:
        raise ValueError("runs_per_scenario must be >= 1")

    ranked_runs = sorted(
        scenario_runs.items(),
        key=lambda kv: (len(kv[1]), kv[0]),
        reverse=True,
    )
    if len(ranked_runs) < runs_per_scenario:
        raise ValueError(
            f"Scenario has only {len(ranked_runs)} runs, but {runs_per_scenario} requested."
        )

    required = []
    required_set = set(required_run_tags or [])
    for run_tag, rows in ranked_runs:
        if run_tag in required_set:
            required.append((run_tag, rows))

    if len(required) > runs_per_scenario:
        runs_per_scenario = len(required)

    selected: List[Tuple[str, List[EpisodeRecord]]] = list(required)
    for run_tag, rows in ranked_runs:
        if run_tag in required_set:
            continue
        if len(selected) >= runs_per_scenario:
            break
        selected.append((run_tag, rows))

    return {run_tag: rows for run_tag, rows in selected}


def _hydrate_rows_from_records(records: Sequence[EpisodeRecord], complete: bool, source: str) -> List[ScheduleRow]:
    out: List[ScheduleRow] = []
    for rec in records:
        out.append(
            ScheduleRow(
                scenario_id=rec.scenario_id,
                run_tag=rec.run_tag,
                episode_index=rec.episode_index,
                mission_name=rec.mission_name,
                episode_id=rec.episode_id,
                episode_dir=rec.episode_dir,
                video_path=rec.video_path,
                survey_path=rec.survey_path,
                complete=complete,
                source=source,
            )
        )
    return out


def _refresh_completed_metadata(
    completed_rows: Sequence[ScheduleRow],
    key_to_record: Dict[Tuple[str, str, int], EpisodeRecord],
) -> List[ScheduleRow]:
    hydrated: List[ScheduleRow] = []
    for row in completed_rows:
        rec = key_to_record.get(row.key)
        if rec is None:
            continue
        hydrated.append(
            ScheduleRow(
                scenario_id=rec.scenario_id,
                run_tag=rec.run_tag,
                episode_index=rec.episode_index,
                mission_name=rec.mission_name,
                episode_id=rec.episode_id,
                episode_dir=rec.episode_dir,
                video_path=rec.video_path,
                survey_path=rec.survey_path,
                complete=True,
                source=row.source,
            )
        )
    return hydrated


def _order_remaining_rows(
    completed_rows: Sequence[ScheduleRow],
    pending_rows: Sequence[ScheduleRow],
    per_scenario_target: int,
    selected_scenarios: Sequence[str],
    rng: random.Random,
) -> List[ScheduleRow]:
    current_counts = {s: 0 for s in selected_scenarios}
    for row in completed_rows:
        current_counts[row.scenario_id] = current_counts.get(row.scenario_id, 0) + 1

    buckets: Dict[str, List[ScheduleRow]] = {s: [] for s in selected_scenarios}
    for row in pending_rows:
        buckets[row.scenario_id].append(row)

    for s in selected_scenarios:
        rng.shuffle(buckets[s])

    ordered: List[ScheduleRow] = []
    prev_scenario = completed_rows[-1].scenario_id if completed_rows else ""

    total_pending = sum(len(v) for v in buckets.values())
    for _ in range(total_pending):
        candidates = [s for s in selected_scenarios if buckets[s]]
        if not candidates:
            break

        min_count = min(current_counts[s] for s in candidates)
        tied = [s for s in candidates if current_counts[s] == min_count]

        non_repeat = [s for s in tied if s != prev_scenario]
        if non_repeat:
            tied = non_repeat

        max_deficit = max(per_scenario_target - current_counts[s] for s in tied)
        deficit_tied = [s for s in tied if (per_scenario_target - current_counts[s]) == max_deficit]

        scenario_pick = rng.choice(deficit_tied)
        row = buckets[scenario_pick].pop()
        ordered.append(row)
        current_counts[scenario_pick] += 1
        prev_scenario = scenario_pick

    return ordered


def _style_header(ws, row_idx: int, columns: int) -> None:
    for col in range(1, columns + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _style_grid(ws, start_row: int, end_row: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _write_workbook(
    rows: Sequence[ScheduleRow],
    output_path: Path,
    selected_scenarios: Sequence[str],
    per_scenario_target: int,
    seed: int,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "human_eval_schedule"

    ws.merge_cells("A1:L1")
    ws["A1"] = "Balanced Human Evaluation Schedule"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = TITLE_FILL

    ws.merge_cells("A2:L3")
    ws["A2"] = (
        f"3 scenarios x {per_scenario_target} evals each ({len(rows)} total). "
        f"Episodes are sampled from indices 2..21 only. Seed={seed}."
    )
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells("A4:L4")
    ws["A4"] = "Selected scenarios: " + " | ".join(selected_scenarios)

    headers = [
        "row_index",
        "complete",
        "video_link",
        "survey_link",
        "notes",
        "scenario_id",
        "run_tag",
        "episode_index",
        "mission_name",
        "episode_id",
        "episode_dir",
        "source",
    ]

    header_row = 6
    for idx, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=text)
    _style_header(ws, header_row, len(headers))

    r = header_row + 1
    for i, row in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=bool(row.complete))
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value=row.scenario_id)
        ws.cell(row=r, column=7, value=row.run_tag)
        ws.cell(row=r, column=8, value=row.episode_index)
        ws.cell(row=r, column=9, value=row.mission_name)
        ws.cell(row=r, column=10, value=row.episode_id)
        ws.cell(row=r, column=11, value=str(row.episode_dir))

        video_cell = ws.cell(row=r, column=3)
        if row.video_path is not None:
            video_cell.value = "video"
            video_cell.hyperlink = row.video_path.as_uri()
            video_cell.font = Font(color="0563C1", underline="single")
        else:
            video_cell.value = ""

        survey_cell = ws.cell(row=r, column=4)
        if row.survey_path is not None:
            survey_cell.value = "survey"
            survey_cell.hyperlink = row.survey_path.as_uri()
            survey_cell.font = Font(color="0563C1", underline="single")
        else:
            survey_cell.value = ""

        ws.cell(row=r, column=12, value=row.source)
        r += 1

    data_end = max(header_row + 1, r - 1)
    _style_grid(ws, header_row, data_end, len(headers))

    widths = {
        1: 10,
        2: 10,
        3: 14,
        4: 14,
        5: 20,
        6: 62,
        7: 20,
        8: 14,
        9: 44,
        10: 24,
        11: 82,
        12: 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:L{data_end}"

    # Summary sheet for quick planning and coverage checks.
    summary_ws = wb.create_sheet("schedule_summary")
    summary_ws.merge_cells("A1:D1")
    summary_ws["A1"] = "Schedule Summary Statistics"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A1"].fill = TITLE_FILL

    summary_ws["A3"] = "Metric"
    summary_ws["B3"] = "Value"
    _style_header(summary_ws, 3, 2)

    total_rows = len(rows)
    total_complete = sum(1 for row in rows if row.complete)
    total_pending = total_rows - total_complete
    total_video_links = sum(1 for row in rows if row.video_path is not None)
    total_survey_links = sum(1 for row in rows if row.survey_path is not None)
    distinct_runs = len({(row.scenario_id, row.run_tag) for row in rows})

    core_metrics = [
        ("total_rows", total_rows),
        ("total_completed", total_complete),
        ("total_pending", total_pending),
        ("total_video_links", total_video_links),
        ("total_survey_links", total_survey_links),
        ("selected_scenarios", len(selected_scenarios)),
        ("distinct_scenario_run_pairs", distinct_runs),
        ("seed", seed),
    ]

    row_ptr = 4
    for metric, value in core_metrics:
        summary_ws.cell(row=row_ptr, column=1, value=metric)
        summary_ws.cell(row=row_ptr, column=2, value=value)
        row_ptr += 1

    summary_ws.cell(row=row_ptr + 1, column=1, value="Per-scenario counts")
    summary_ws.cell(row=row_ptr + 2, column=1, value="scenario_id")
    summary_ws.cell(row=row_ptr + 2, column=2, value="rows")
    summary_ws.cell(row=row_ptr + 2, column=3, value="completed")
    summary_ws.cell(row=row_ptr + 2, column=4, value="pending")
    _style_header(summary_ws, row_ptr + 2, 4)

    scenario_counts: Dict[str, int] = {s: 0 for s in selected_scenarios}
    scenario_completed: Dict[str, int] = {s: 0 for s in selected_scenarios}
    for row in rows:
        scenario_counts[row.scenario_id] = scenario_counts.get(row.scenario_id, 0) + 1
        if row.complete:
            scenario_completed[row.scenario_id] = scenario_completed.get(row.scenario_id, 0) + 1

    row_ptr = row_ptr + 3
    for scenario in selected_scenarios:
        count = scenario_counts.get(scenario, 0)
        done = scenario_completed.get(scenario, 0)
        summary_ws.cell(row=row_ptr, column=1, value=scenario)
        summary_ws.cell(row=row_ptr, column=2, value=count)
        summary_ws.cell(row=row_ptr, column=3, value=done)
        summary_ws.cell(row=row_ptr, column=4, value=count - done)
        row_ptr += 1

    row_ptr += 1
    summary_ws.cell(row=row_ptr, column=1, value="Per-run distribution")
    summary_ws.cell(row=row_ptr + 1, column=1, value="scenario_id")
    summary_ws.cell(row=row_ptr + 1, column=2, value="run_tag")
    summary_ws.cell(row=row_ptr + 1, column=3, value="rows")
    summary_ws.cell(row=row_ptr + 1, column=4, value="completed")
    _style_header(summary_ws, row_ptr + 1, 4)

    run_counts: Dict[Tuple[str, str], int] = {}
    run_completed: Dict[Tuple[str, str], int] = {}
    for row in rows:
        key = (row.scenario_id, row.run_tag)
        run_counts[key] = run_counts.get(key, 0) + 1
        if row.complete:
            run_completed[key] = run_completed.get(key, 0) + 1

    row_ptr = row_ptr + 2
    for scenario in selected_scenarios:
        run_tags = sorted({row.run_tag for row in rows if row.scenario_id == scenario})
        for run_tag in run_tags:
            key = (scenario, run_tag)
            summary_ws.cell(row=row_ptr, column=1, value=scenario)
            summary_ws.cell(row=row_ptr, column=2, value=run_tag)
            summary_ws.cell(row=row_ptr, column=3, value=run_counts.get(key, 0))
            summary_ws.cell(row=row_ptr, column=4, value=run_completed.get(key, 0))
            row_ptr += 1

    for r in range(3, row_ptr):
        for c in range(1, 5):
            cell = summary_ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    summary_ws.column_dimensions["A"].width = 72
    summary_ws.column_dimensions["B"].width = 24
    summary_ws.column_dimensions["C"].width = 14
    summary_ws.column_dimensions["D"].width = 14
    summary_ws.freeze_panes = "A4"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_schedule(
    logs_dir: Path,
    selected_scenarios: Sequence[str],
    per_scenario_target: int,
    runs_per_scenario: int,
    seed: int,
    existing_workbook: Optional[Path],
    output_path: Path,
) -> Dict[str, object]:
    rng = random.Random(seed)

    all_records = _load_summary_records(logs_dir)
    grouped = _group_by_scenario_run(all_records)

    records = [r for r in all_records if r.scenario_id in set(selected_scenarios)]
    key_to_record = {r.key: r for r in records}

    auto_completed = _load_episode_workbook_completions(
        records=records,
        selected_scenarios=selected_scenarios,
    )

    completed_prior = _load_completed_rows(
        existing_workbook=existing_workbook,
        selected_scenarios=selected_scenarios,
        valid_keys=key_to_record.keys(),
    )

    # Merge existing-workbook completions and per-episode workbook completions.
    merged_completed: Dict[Tuple[str, str, int], ScheduleRow] = {}
    for row in completed_prior:
        merged_completed[row.key] = row
    for row in auto_completed:
        if row.key not in merged_completed:
            merged_completed[row.key] = row

    completed_rows = _refresh_completed_metadata(list(merged_completed.values()), key_to_record)

    scenario_run_maps: Dict[str, Dict[str, List[EpisodeRecord]]] = {}
    for scenario in selected_scenarios:
        if scenario not in grouped:
            raise ValueError(f"Selected scenario missing: {scenario}")
        required_runs = sorted({row.run_tag for row in completed_rows if row.scenario_id == scenario})
        scenario_run_maps[scenario] = _select_runs_for_scenario(
            grouped[scenario],
            runs_per_scenario=runs_per_scenario,
            required_run_tags=required_runs,
        )

    # Keep completed rows at the top, but interleave scenarios where possible.
    completed_rows = _order_remaining_rows(
        completed_rows=[],
        pending_rows=completed_rows,
        per_scenario_target=per_scenario_target,
        selected_scenarios=selected_scenarios,
        rng=rng,
    )
    for row in completed_rows:
        row.complete = True

    completed_by_scenario: Dict[str, List[ScheduleRow]] = {s: [] for s in selected_scenarios}
    for row in completed_rows:
        completed_by_scenario[row.scenario_id].append(row)

    pending_records: List[EpisodeRecord] = []
    used_keys = {row.key for row in completed_rows}

    for scenario in selected_scenarios:
        completed_count = len(completed_by_scenario[scenario])
        if completed_count > per_scenario_target:
            raise ValueError(
                f"Scenario {scenario} already has {completed_count} completed rows, "
                f"which exceeds target {per_scenario_target}."
            )

        need = per_scenario_target - completed_count
        if need == 0:
            continue

        picks = _pick_rows_for_scenario(
            scenario_runs=scenario_run_maps[scenario],
            total_needed=need,
            completed_keys=used_keys,
            rng=rng,
        )
        for rec in picks:
            used_keys.add(rec.key)
        pending_records.extend(picks)

    pending_rows = _hydrate_rows_from_records(pending_records, complete=False, source="new")

    ordered_pending = _order_remaining_rows(
        completed_rows=completed_rows,
        pending_rows=pending_rows,
        per_scenario_target=per_scenario_target,
        selected_scenarios=selected_scenarios,
        rng=rng,
    )

    final_rows = list(completed_rows) + ordered_pending
    expected_total = per_scenario_target * len(selected_scenarios)
    if len(final_rows) != expected_total:
        raise ValueError(f"Built {len(final_rows)} rows, expected {expected_total}.")

    scenario_totals: Dict[str, int] = {s: 0 for s in selected_scenarios}
    run_totals: Dict[str, Dict[str, int]] = {s: {} for s in selected_scenarios}
    completed_total = 0

    for row in final_rows:
        scenario_totals[row.scenario_id] += 1
        run_totals[row.scenario_id][row.run_tag] = run_totals[row.scenario_id].get(row.run_tag, 0) + 1
        if row.complete:
            completed_total += 1

    _write_workbook(
        rows=final_rows,
        output_path=output_path,
        selected_scenarios=selected_scenarios,
        per_scenario_target=per_scenario_target,
        seed=seed,
    )

    return {
        "ok": True,
        "output_path": str(output_path),
        "selected_scenarios": list(selected_scenarios),
        "rows": len(final_rows),
        "per_scenario_target": per_scenario_target,
        "runs_per_scenario": runs_per_scenario,
        "completed_rows_carried_over": completed_total,
        "completed_rows_detected_from_episode_workbooks": len(auto_completed),
        "scenario_totals": scenario_totals,
        "run_totals": run_totals,
        "existing_workbook": str(existing_workbook) if existing_workbook else "",
        "seed": seed,
    }


def main() -> None:
    args = _parse_args()

    project_root = Path(args.project_root).expanduser().resolve()
    logs_dir = Path(args.logs_dir).expanduser().resolve() if args.logs_dir else (project_root / "data" / "logs")
    if not logs_dir.exists():
        raise SystemExit(f"Logs directory not found: {logs_dir}")

    grouped = _group_by_scenario_run(_load_summary_records(logs_dir))
    selected_scenarios = _choose_scenarios(grouped, args.scenarios)

    existing_path: Optional[Path] = None
    if args.existing_workbook:
        existing_path = Path(args.existing_workbook).expanduser().resolve()

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        stamp = dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        output_path = logs_dir / f"human_eval_schedule_{stamp}.xlsx"

    result = generate_schedule(
        logs_dir=logs_dir,
        selected_scenarios=selected_scenarios,
        per_scenario_target=args.evals_per_scenario,
        runs_per_scenario=args.runs_per_scenario,
        seed=args.seed,
        existing_workbook=existing_path,
        output_path=output_path,
    )
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
